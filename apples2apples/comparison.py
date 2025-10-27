import pandas as pd
import numpy as np
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import re
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def extract_product_type(name):
    """Extract the core product type, removing brands, sizes, and specific details"""
    if pd.isna(name):
        return ""
    
    name = str(name).lower()
    
    # Remove brand names (extensive list)
    brands_to_remove = [
        'groupr', 'nathan\'s', 'jimmy dean', 'carolina', 'rao\'s', 'butterball',
        'kellogg\'s', 'hellmann\'s', 'goya', 'bubba', 'wonder', 'libby\'s',
        'badia', 'cheerios', 'velveeta', 'oreo', 'simply', 'martin\'s',
        'cinnamon toast crunch', 'bell & evans', 'blue diamond', 'mott\'s',
        'chips ahoy', 'lactaid', 'froot loops', 'stonyfield', 'sara lee',
        'mission', 'entenmann\'s', 'boar\'s head', 'specially selected',
        'kirkwood', 'happy farms', 'nature\'s nectar', 'benton\'s', 
        'clancy\'s', 'simply nature', 'l\'oven fresh', 'park street deli',
        'reggano', 'fresh', 'organic', 'premium', 'classic', 'original',
        'natural', 'pure', 'real', 'homemade', 'famous', 'best',
        'perdue', 'tyson', 'foster farms', 'oscar mayer', 'hillshire farm',
        'kraft', 'philadelphia', 'tropicana', 'minute maid', 'coca-cola',
        'pepsi', 'poland spring', 'dasani', 'aquafina', 'smart water',
        'tide', 'gain', 'downy', 'charmin', 'bounty', 'dawn', 'lysol',
        'clorox', 'arm & hammer', 'colgate', 'crest', 'listerine',
        'pantene', 'head & shoulders', 'dove', 'olay', 'nivea',
        'general mills', 'post', 'quaker', 'kashi', 'nature valley',
        'planters', 'lay\'s', 'doritos', 'cheetos', 'pringles',
        'nabisco', 'pepperidge farm', 'keebler', 'sunshine', 'ritz',
        'wheat thins', 'triscuit', 'goldfish', 'campbell\'s', 'progresso',
        'hunt\'s', 'del monte', 'dole', 'chiquita'
    ]
    
    for brand in brands_to_remove:
        name = name.replace(brand, '').strip()
    
    # Remove size and quantity information
    name = re.sub(r'\d+\.?\d*\s*(oz|lb|ct|count|pk|pack|gallon|btl|bottle|jar|pkg|package|jug|liter|fl oz|quart|pint)', '', name)
    name = re.sub(r'\d+\s*(pack|ct|count|piece|pieces)', '', name)
    name = re.sub(r'\([^)]*\)', '', name)  # Remove anything in parentheses
    name = re.sub(r'\d+\s*-\s*\d+', '', name)  # Remove ranges like "4-6"
    name = re.sub(r'apx\s*\d+\.?\d*', '', name)  # Remove "apx 1.25"
    name = re.sub(r'\$\d+\.?\d*', '', name)  # Remove prices
    
    # Remove descriptive words that don't help with matching
    descriptors_to_remove = [
        'sliced', 'diced', 'chopped', 'whole', 'half', 'quarter',
        'thin', 'thick', 'large', 'small', 'medium', 'jumbo', 'mini',
        'frozen', 'fresh', 'canned', 'jarred', 'bottled', 'packed',
        'boneless', 'skinless', 'bone-in', 'skin-on', 'seedless',
        'unsalted', 'salted', 'low fat', 'fat free', 'sugar free',
        'gluten free', 'organic', 'natural', 'free range', 'cage free',
        'antibiotic free', 'hormone free', 'grass fed', 'wild caught',
        'farm raised', 'center cut', 'family pack', 'value pack',
        'super pack', 'mega roll', 'double roll', 'ultra', 'extra',
        'premium', 'select', 'choice', 'prime', 'grade a', 'usda',
        'certified', 'all natural', 'no added', 'reduced', 'light',
        'diet', 'zero', 'caffeine free', 'decaf', 'regular'
    ]
    
    for descriptor in descriptors_to_remove:
        name = name.replace(descriptor, '').strip()
    
    # Clean up extra spaces and punctuation
    name = re.sub(r'[^\w\s]', ' ', name)  # Replace punctuation with spaces
    name = ' '.join(name.split())  # Remove extra spaces
    
    # Map common variations to standard terms
    product_mappings = {
        # Meat products
        'beef franks': 'hot dogs',
        'skinless beef franks': 'hot dogs',
        'bun length franks': 'hot dogs',
        'beef hot dogs': 'hot dogs',
        'wieners': 'hot dogs',
        'frankfurters': 'hot dogs',
        
        # Bread products
        'hamburger buns': 'burger buns',
        'sandwich buns': 'burger buns',
        'hot dog rolls': 'hot dog buns',
        'buns': 'burger buns',
        
        # Eggs
        'large eggs': 'eggs',
        'white eggs': 'eggs',
        'brown eggs': 'eggs',
        
        # Bacon
        'applewood smoked bacon': 'bacon',
        'hickory smoked bacon': 'bacon',
        'thick cut bacon': 'bacon',
        
        # Ground meat
        'ground turkey': 'turkey ground',
        'ground beef': 'beef ground',
        'ground chicken': 'chicken ground',
        
        # Rice
        'jasmine rice': 'rice',
        'long grain rice': 'rice',
        'white rice': 'rice',
        'brown rice': 'rice',
        
        # Pasta sauce
        'marinara sauce': 'pasta sauce',
        'tomato sauce': 'pasta sauce',
        'spaghetti sauce': 'pasta sauce',
        
        # Cereal
        'frosted flakes': 'cereal',
        'corn flakes': 'cereal',
        'cinnamon toast crunch': 'cereal',
        'cheerios': 'cereal',
        'froot loops': 'cereal',
        
        # Cheese
        'american cheese': 'cheese',
        'cheddar cheese': 'cheese',
        'swiss cheese': 'cheese',
        'provolone cheese': 'cheese',
        'mozzarella cheese': 'cheese',
        
        # Milk
        'whole milk': 'milk',
        'skim milk': 'milk',
        '2% milk': 'milk',
        '1% milk': 'milk',
        'almond milk': 'almond milk',
        
        # Juice
        'orange juice': 'orange juice',
        'apple juice': 'apple juice',
        'grape juice': 'grape juice',
        'cranberry juice': 'cranberry juice',
        
        # Water
        'purified water': 'water',
        'spring water': 'water',
        'drinking water': 'water',
        'bottled water': 'water'
    }
    
    # Apply mappings
    for variation, standard in product_mappings.items():
        if variation in name:
            name = standard
            break
    
    return name.strip()

def extract_price(price_str):
    """Extract numeric price from price string"""
    if pd.isna(price_str):
        return None
    
    price_str = str(price_str).replace('$', '').replace(',', '')
    
    # Handle "Buy X Get Y Free" patterns
    if 'buy' in price_str.lower() and 'get' in price_str.lower() and 'free' in price_str.lower():
        return None  # Skip promotional pricing for now
    
    # Handle ranges like "2 for $3.00"
    if 'for $' in price_str:
        parts = price_str.split('for $')
        if len(parts) == 2:
            try:
                quantity = float(parts[0].strip())
                total_price = float(parts[1].strip())
                return total_price / quantity
            except:
                pass
    
    # Handle "3 for $X" pattern
    for_pattern = re.search(r'(\d+)\s*for\s*\$?(\d+\.?\d*)', price_str)
    if for_pattern:
        try:
            quantity = float(for_pattern.group(1))
            total_price = float(for_pattern.group(2))
            return total_price / quantity
        except:
            pass
    
    # Extract first number that looks like a price
    match = re.search(r'(\d+\.?\d*)', price_str)
    if match:
        try:
            return float(match.group(1))
        except:
            return None
    
    return None

def find_best_match(product_type, store_products, threshold=70):
    """Find the best matching product type in store inventory"""
    if not product_type:
        return None, 0
    
    # Create a list of product types from store products
    store_product_types = [extract_product_type(p) for p in store_products]
    
    # Find best match using fuzzy string matching
    best_match = process.extractOne(product_type, store_product_types, 
                                   scorer=fuzz.token_sort_ratio)
    
    if best_match and best_match[1] >= threshold:
        # Find the original product name
        original_index = store_product_types.index(best_match[0])
        return store_products[original_index], best_match[1]
    
    # Try with lower threshold for partial matches
    if threshold > 60:
        return find_best_match(product_type, store_products, threshold=60)
    
    return None, 0

def load_and_process_stores():
    """Load all store data"""
    stores = {}
    
    # Load Groupr (reference store)
    groupr_df = pd.read_csv('Castle Hill - Groupr.csv')
    stores['Groupr'] = {
        'df': groupr_df,
        'name_col': 'Product Name',
        'price_col': 'Price'
    }
    
    # Load Aldi
    aldi_df = pd.read_csv('Castle Hill - Aldi.csv')
    stores['Aldi'] = {
        'df': aldi_df,
        'name_col': 'name',
        'price_col': 'price'
    }
    
    # Load KeyFoods
    keyfoods_df = pd.read_csv('Castle Hill - KeyFoods.csv')
    stores['KeyFoods'] = {
        'df': keyfoods_df,
        'name_col': 'Product Name',
        'price_col': 'Price'
    }
    
    # Load ShopRite
    shoprite_df = pd.read_csv('Castle Hill - ShopRite.csv')
    stores['ShopRite'] = {
        'df': shoprite_df,
        'name_col': 'Product Name',
        'price_col': 'Price'
    }
    
    return stores

def create_price_comparison():
    """Create comprehensive price comparison based on product types"""
    stores = load_and_process_stores()
    
    # Get Groupr products as reference
    groupr_products = stores['Groupr']['df']['Product Name'].tolist()
    groupr_prices = stores['Groupr']['df']['Price'].tolist()
    
    # Initialize results
    results = []
    
    print("Processing products by type...")
    
    for i, product in enumerate(groupr_products):
        product_type = extract_product_type(product)
        print(f"Processing: {product} -> Type: '{product_type}'")
        
        result_row = {
            'Original_Product_Name': product,
            'Product_Type': product_type,
            'Groupr_Price': extract_price(groupr_prices[i])
        }
        
        # Find matches in other stores
        for store_name, store_info in stores.items():
            if store_name == 'Groupr':
                continue
                
            store_df = store_info['df']
            name_col = store_info['name_col']
            price_col = store_info['price_col']
            
            # Get store product names
            store_products = store_df[name_col].fillna('').tolist()
            
            # Find best match
            best_match, confidence = find_best_match(product_type, store_products)
            
            if best_match and confidence >= 60:
                # Get the price for the matched product
                matched_row = store_df[store_df[name_col] == best_match].iloc[0]
                price = extract_price(matched_row[price_col])
                
                result_row[f'{store_name}_Product'] = best_match
                result_row[f'{store_name}_Price'] = price
                result_row[f'{store_name}_Confidence'] = confidence
                
                print(f"  - Found in {store_name}: {best_match} (${price}) [Confidence: {confidence}%]")
            else:
                result_row[f'{store_name}_Product'] = 'Not Found'
                result_row[f'{store_name}_Price'] = None
                result_row[f'{store_name}_Confidence'] = 0
                print(f"  - Not found in {store_name}")
        
        results.append(result_row)
        print()
    
    return results

def save_to_excel_with_highlighting(results, filename='price_comparison_by_type.xlsx'):
    """Save results to Excel with lowest price highlighting"""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Price Comparison by Type"
    
    # Define headers
    headers = ['Original_Product_Name', 'Product_Type', 'Groupr_Price', 
               'Aldi_Product', 'Aldi_Price', 'Aldi_Confidence',
               'KeyFoods_Product', 'KeyFoods_Price', 'KeyFoods_Confidence',
               'ShopRite_Product', 'ShopRite_Price', 'ShopRite_Confidence',
               'Lowest_Price_Store', 'Savings_vs_Groupr']
    
    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Define fill colors
    lowest_price_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
    savings_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # Gold
    
    # Write data
    for row_idx, result in enumerate(results, start=2):
        # Basic data
        ws.cell(row=row_idx, column=1, value=result['Original_Product_Name'])
        ws.cell(row=row_idx, column=2, value=result['Product_Type'])
        ws.cell(row=row_idx, column=3, value=result['Groupr_Price'])
        
        col_idx = 4
        for store in ['Aldi', 'KeyFoods', 'ShopRite']:
            ws.cell(row=row_idx, column=col_idx, value=result.get(f'{store}_Product', 'Not Found'))
            ws.cell(row=row_idx, column=col_idx+1, value=result.get(f'{store}_Price'))
            ws.cell(row=row_idx, column=col_idx+2, value=result.get(f'{store}_Confidence', 0))
            col_idx += 3
        
        # Find lowest price and highlight
        prices = {}
        if result['Groupr_Price'] is not None:
            prices['Groupr'] = (result['Groupr_Price'], 3)  # column 3
        
        for store, col in [('Aldi', 5), ('KeyFoods', 8), ('ShopRite', 11)]:
            price = result.get(f'{store}_Price')
            if price is not None and price > 0:
                prices[store] = (price, col)
        
        if prices:
            lowest_store = min(prices.keys(), key=lambda x: prices[x][0])
            lowest_price = prices[lowest_store][0]
            lowest_col = prices[lowest_store][1]
            
            # Highlight the lowest price cell
            ws.cell(row=row_idx, column=lowest_col).fill = lowest_price_fill
            
            # Add lowest price store info
            ws.cell(row=row_idx, column=13, value=f"{lowest_store} (${lowest_price:.2f})")
            
            # Calculate savings vs Groupr
            groupr_price = result['Groupr_Price']
            if groupr_price is not None and lowest_price < groupr_price:
                savings = groupr_price - lowest_price
                savings_pct = (savings / groupr_price) * 100
                savings_cell = ws.cell(row=row_idx, column=14, value=f"${savings:.2f} ({savings_pct:.1f}%)")
                if savings > 0:
                    savings_cell.fill = savings_fill
        else:
            ws.cell(row=row_idx, column=13, value="No comparable prices found")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(filename)
    print(f"Results saved to {filename}")

def save_to_csv(results, filename='price_comparison_by_type.csv'):
    """Save results to CSV format as well"""
    
    # Prepare data for CSV
    csv_data = []
    
    for result in results:
        row = {
            'Original_Product_Name': result['Original_Product_Name'],
            'Product_Type': result['Product_Type'],
            'Groupr_Price': result['Groupr_Price'],
            'Aldi_Product': result.get('Aldi_Product', 'Not Found'),
            'Aldi_Price': result.get('Aldi_Price'),
            'Aldi_Confidence': result.get('Aldi_Confidence', 0),
            'KeyFoods_Product': result.get('KeyFoods_Product', 'Not Found'),
            'KeyFoods_Price': result.get('KeyFoods_Price'),
            'KeyFoods_Confidence': result.get('KeyFoods_Confidence', 0),
            'ShopRite_Product': result.get('ShopRite_Product', 'Not Found'),
            'ShopRite_Price': result.get('ShopRite_Price'),
            'ShopRite_Confidence': result.get('ShopRite_Confidence', 0)
        }
        
        # Find lowest price
        prices = {}
        if result['Groupr_Price'] is not None:
            prices['Groupr'] = result['Groupr_Price']
        
        for store in ['Aldi', 'KeyFoods', 'ShopRite']:
            price = result.get(f'{store}_Price')
            if price is not None and price > 0:
                prices[store] = price
        
        if prices:
            lowest_store = min(prices.keys(), key=lambda x: prices[x])
            lowest_price = prices[lowest_store]
            row['Lowest_Price_Store'] = f"{lowest_store} (${lowest_price:.2f})"
            
            # Calculate savings vs Groupr
            groupr_price = result['Groupr_Price']
            if groupr_price is not None and lowest_price < groupr_price:
                savings = groupr_price - lowest_price
                savings_pct = (savings / groupr_price) * 100
                row['Savings_vs_Groupr'] = f"${savings:.2f} ({savings_pct:.1f}%)"
            else:
                row['Savings_vs_Groupr'] = "No savings"
        else:
            row['Lowest_Price_Store'] = 'No prices found'
            row['Savings_vs_Groupr'] = 'No comparison'
        
        csv_data.append(row)
    
    # Write to CSV
    fieldnames = ['Original_Product_Name', 'Product_Type', 'Groupr_Price', 
                 'Aldi_Product', 'Aldi_Price', 'Aldi_Confidence',
                 'KeyFoods_Product', 'KeyFoods_Price', 'KeyFoods_Confidence',
                 'ShopRite_Product', 'ShopRite_Price', 'ShopRite_Confidence',
                 'Lowest_Price_Store', 'Savings_vs_Groupr']
    
    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(csv_data)
    
    print(f"CSV results saved to {filename}")

def main():
    """Main execution function"""
    print("Starting price comparison analysis by product type...")
    print("This approach focuses on generic product categories rather than specific brands...")
    
    try:
        # Create price comparison
        results = create_price_comparison()
        
        # Save results
        save_to_excel_with_highlighting(results)
        save_to_csv(results)
        
        # Print summary
        print(f"\nSummary:")
        print(f"Total products processed: {len(results)}")
        
        found_counts = {}
        total_savings = 0
        products_with_savings = 0
        
        for store in ['Aldi', 'KeyFoods', 'ShopRite']:
            found_count = sum(1 for r in results if r.get(f'{store}_Price') is not None)
            found_counts[store] = found_count
            print(f"Products found in {store}: {found_count}")
        
        # Calculate potential savings
        for result in results:
            groupr_price = result['Groupr_Price']
            if groupr_price is not None:
                competitor_prices = []
                for store in ['Aldi', 'KeyFoods', 'ShopRite']:
                    price = result.get(f'{store}_Price')
                    if price is not None and price > 0:
                        competitor_prices.append(price)
                
                if competitor_prices:
                    lowest_competitor = min(competitor_prices)
                    if lowest_competitor < groupr_price:
                        total_savings += (groupr_price - lowest_competitor)
                        products_with_savings += 1
        
        print(f"\nPotential Savings Analysis:")
        print(f"Products with better prices elsewhere: {products_with_savings}")
        print(f"Total potential savings: ${total_savings:.2f}")
        if products_with_savings > 0:
            print(f"Average savings per product: ${total_savings/products_with_savings:.2f}")
        
        print("\nFiles created:")
        print("- price_comparison_by_type.xlsx (with highlighting and savings analysis)")
        print("- price_comparison_by_type.csv (plain CSV)")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()